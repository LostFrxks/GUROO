from datetime import datetime, timedelta
from pathlib import Path
from uuid import uuid4
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status, Request
from pydantic import BaseModel
from fastapi.responses import FileResponse
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user, get_optional_user
from app.core.config import settings
from app.core.time import now as time_now
from app.db.session import get_db
from app.models.tutor import Notification, TutorProfile, TutorStudent
from app.models.user import User, UserRole
from app.schemas.tutor import (
    AttendanceSubmit,
    LessonDetailsUpdate,
    LessonAttendanceUpdate,
    TutorListItem,
    TutorProfileOut,
    TutorProfileUpdateRequest,
    TutorRequestCreate,
    TutorScheduleResponse,
)

router = APIRouter()

LESSON_TIMES = {
    "1": "08:00-09:20",
    "2": "09:30-10:50",
    "3": "11:40-13:00",
    "4": "13:10-14:30",
    "5": "14:40-16:00",
    "6": "16:00-17:20",
    "7": "17:30-19:00",
}

DAY_MAP = {
    "Monday": "Monday",
    "Tuesday": "Tuesday",
    "Wednesday": "Wednesday",
    "Thursday": "Thursday",
    "Friday": "Friday",
    "Saturday": "Saturday",
    "Sunday": "Sunday",
}

LESSON_DIR = Path(settings.media_root) / "lessons"
LESSON_DIR.mkdir(parents=True, exist_ok=True)

RU_DAY_INDEX = {
    "Понедельник": 0,
    "Вторник": 1,
    "Среда": 2,
    "Четверг": 3,
    "Пятница": 4,
    "Суббота": 5,
    "Воскресенье": 6,
}


def _normalize_schedule_entries(schedule: list | dict) -> list[dict]:
    entries: list[dict] = []
    if isinstance(schedule, list):
        for entry in schedule:
            day = str(entry.get("day") or "").strip()
            pair = str(entry.get("pair") or "").strip()
            if not day or not pair:
                continue
            entries.append(
                {
                    "day": day,
                    "pair": pair,
                    "location": str(entry.get("location") or "").strip(),
                }
            )
    elif isinstance(schedule, dict):
        for day, pairs in schedule.items():
            for pair in pairs or []:
                entries.append({"day": str(day), "pair": str(pair), "location": ""})
    return entries


def _lesson_window(now: datetime, entry: dict) -> tuple[datetime, datetime] | None:
    day = entry.get("day", "")
    pair = entry.get("pair", "")
    if not day or pair not in LESSON_TIMES:
        return None
    weekday = RU_DAY_INDEX.get(day)
    if weekday is None or weekday != now.weekday():
        return None
    start_str, end_str = LESSON_TIMES[pair].split("-")
    start_hour, start_min = map(int, start_str.split(":"))
    end_hour, end_min = map(int, end_str.split(":"))
    start = now.replace(hour=start_hour, minute=start_min, second=0, microsecond=0)
    end = now.replace(hour=end_hour, minute=end_min, second=0, microsecond=0)
    return start, end


async def _resolve_tutor_profile_id(db: AsyncSession, tutor_id: int) -> int | None:
    """
    Some clients may send tutor_id as TutorProfile.id, others as User.id.
    Try both so we can find the relation reliably.
    """
    direct = await db.execute(select(TutorProfile.id).where(TutorProfile.id == tutor_id))
    profile_id = direct.scalar_one_or_none()
    if profile_id is not None:
        return int(profile_id)
    by_user = await db.execute(select(TutorProfile.id).where(TutorProfile.user_id == tutor_id))
    profile_id = by_user.scalar_one_or_none()
    if profile_id is None:
        return None
    return int(profile_id)


@router.get("/", response_model=list[TutorListItem])
async def list_tutors(
    user_id: int | None = None,
    current_user: User | None = Depends(get_optional_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(TutorProfile).options(selectinload(TutorProfile.user)))
    tutors = result.scalars().all()

    confirmed_map: dict[int, bool] = {}
    status_map: dict[int, str | None] = {}
    cooldown_map: dict[int, datetime] = {}
    effective_user_id = user_id or (current_user.id if current_user else None)
    if effective_user_id:
        relations = await db.execute(
            select(TutorStudent).where(TutorStudent.student_id == effective_user_id)
        )
        now = datetime.now(tz=ZoneInfo("Asia/Bishkek"))
        for rel in relations.scalars().all():
            if rel.confirmed:
                confirmed_map[rel.tutor_id] = rel.attended
                status_map[rel.tutor_id] = "confirmed"
            elif rel.declined_until and rel.declined_until > now:
                status_map[rel.tutor_id] = "cooldown"
                cooldown_map[rel.tutor_id] = rel.declined_until
            else:
                status_map[rel.tutor_id] = "pending"

    items: list[TutorListItem] = []
    for tutor in tutors:
        user = tutor.user
        is_incomplete_profile = (
            not (tutor.subject or "").strip()
            or not (tutor.bio or "").strip()
            or not (tutor.phone or "").strip()
            or not (tutor.telegram or "").strip()
            or not tutor.course
            or not tutor.schedule
        )
        if is_incomplete_profile:
            continue
        name = f"{user.first_name or ''} {user.last_name or ''}".strip()
        photo_paths = tutor.photo_paths or ([] if tutor.photo_path is None else [tutor.photo_path])
        items.append(
            TutorListItem(
                id=tutor.id,
                user_id=user.id,
                name=name,
                subject=tutor.subject,
                bio=tutor.bio,
                photo_path=tutor.photo_path,
                photo_paths=photo_paths,
                schedule=tutor.schedule or {},
                course=tutor.course,
                phone=tutor.phone,
                telegram=tutor.telegram,
                group=tutor.group,
                own_course=tutor.own_course,
                request_status=status_map.get(tutor.id),
                attended=confirmed_map.get(tutor.id, False),
                cooldown_until=cooldown_map.get(tutor.id),
            )
        )
    return items


@router.get("", response_model=list[TutorListItem], include_in_schema=False)
async def list_tutors_no_slash(
    user_id: int | None = None,
    current_user: User | None = Depends(get_optional_user),
    db: AsyncSession = Depends(get_db),
):
    return await list_tutors(user_id=user_id, current_user=current_user, db=db)


@router.get("/me", response_model=TutorProfileOut)
async def get_my_profile(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == current_user.id))
    profile = result.scalar_one_or_none()
    if not profile:
        profile = TutorProfile(
            user_id=current_user.id,
            first_name=current_user.first_name,
            last_name=current_user.last_name,
            subject="",
            schedule={},
        )
        db.add(profile)
        await db.commit()
        await db.refresh(profile)
    return profile


@router.post("/me/photos")
async def upload_profile_photos(
    files: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == current_user.id))
    profile = result.scalar_one_or_none()
    if not profile:
        profile = TutorProfile(user_id=current_user.id, schedule={})
        db.add(profile)
        await db.flush()

    if not files or len(files) == 0:
        raise HTTPException(status_code=400, detail="No files provided")
    existing_paths = list(profile.photo_paths or [])
    if len(existing_paths) + len(files) > 3:
        raise HTTPException(status_code=400, detail="Max 3 photos allowed")

    photos_dir = Path(settings.media_root) / "tutors"
    photos_dir.mkdir(parents=True, exist_ok=True)

    saved_paths: list[str] = []
    for file in files:
        suffix = Path(file.filename or "").suffix or ".jpg"
        safe_name = f"{uuid4().hex}{suffix}"
        out_path = photos_dir / safe_name
        content = await file.read()
        out_path.write_bytes(content)
        saved_paths.append(f"/media/tutors/{safe_name}")

    profile.photo_paths = existing_paths + saved_paths
    profile.photo_path = profile.photo_paths[0] if profile.photo_paths else None
    await db.commit()
    await db.refresh(profile)
    return {"photo_paths": profile.photo_paths, "photo_path": profile.photo_path}


class PhotoDeleteRequest(BaseModel):
    path: str


@router.delete("/me/photos")
async def delete_profile_photo(
    payload: PhotoDeleteRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == current_user.id))
    profile = result.scalar_one_or_none()
    if not profile:
        raise HTTPException(status_code=404, detail="Profile not found")

    current_paths = list(profile.photo_paths or [])
    if payload.path in current_paths:
        current_paths.remove(payload.path)
        profile.photo_paths = current_paths
        profile.photo_path = current_paths[0] if current_paths else None
        try:
            if payload.path.startswith("/media/"):
                local_path = Path(settings.media_root) / payload.path.replace("/media/", "")
                if local_path.exists():
                    local_path.unlink()
        except Exception:
            pass
        await db.commit()
        await db.refresh(profile)
    return {"photo_paths": profile.photo_paths or [], "photo_path": profile.photo_path}


@router.put("/me", response_model=TutorProfileOut)
async def update_my_profile(
    payload: TutorProfileUpdateRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == current_user.id))
    profile = result.scalar_one_or_none()
    if not profile:
        profile = TutorProfile(user_id=current_user.id, schedule={})
        db.add(profile)

    data = payload.profile.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(profile, key, value)

    await db.commit()
    await db.refresh(profile)
    return profile


@router.get("/registered")
async def get_registered_tutors(
    user_id: int,
    kind: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    now = datetime.now(tz=ZoneInfo("Asia/Bishkek"))

    def status_for(relation: TutorStudent) -> str:
        if relation.confirmed:
            return "confirmed"
        if relation.declined_until and relation.declined_until > now:
            return "cooldown"
        return "pending"

    if user.role == UserRole.student:
        relations = await db.execute(
            select(TutorStudent, TutorProfile, User)
            .join(TutorProfile, TutorProfile.id == TutorStudent.tutor_id)
            .join(User, User.id == TutorProfile.user_id)
            .where(TutorStudent.student_id == user_id)
            .where(TutorStudent.confirmed == True)  # noqa: E712
        )
        tutors_list = []
        for relation, tutor_profile, tutor_user in relations.all():
            tutors_list.append(
                {
                    "name": f"{tutor_user.first_name} {tutor_user.last_name}",
                    "subject": tutor_profile.subject,
                    "phone": tutor_profile.phone,
                    "telegram": tutor_profile.telegram,
                    "schedule": tutor_profile.schedule,
                    "status": status_for(relation),
                    "cooldown_until": relation.declined_until,
                }
            )
        if kind == "students":
            return []
        return tutors_list

    if user.role == UserRole.tutor:
        profile_result = await db.execute(
            select(TutorProfile.id).where(TutorProfile.user_id == user_id)
        )
        tutor_profile_id = profile_result.scalar_one_or_none()
        students_list = []
        if tutor_profile_id:
            students_rel = await db.execute(
                select(TutorStudent, User)
                .join(User, User.id == TutorStudent.student_id)
                .where(TutorStudent.tutor_id == tutor_profile_id)
            )
            for relation, student_user in students_rel.all():
                students_list.append(
                    {
                        "name": f"{student_user.first_name} {student_user.last_name}",
                        "email": student_user.email,
                        "status": status_for(relation),
                        "cooldown_until": relation.declined_until,
                        "confirmed": relation.confirmed,
                    }
                )

        tutors_rel = await db.execute(
            select(TutorStudent, TutorProfile, User)
            .join(TutorProfile, TutorProfile.id == TutorStudent.tutor_id)
            .join(User, User.id == TutorProfile.user_id)
            .where(TutorStudent.student_id == user_id)
            .where(TutorStudent.confirmed == True)  # noqa: E712
        )
        tutors_list = []
        for relation, tutor_profile, tutor_user in tutors_rel.all():
            tutors_list.append(
                {
                    "name": f"{tutor_user.first_name} {tutor_user.last_name}",
                    "subject": tutor_profile.subject,
                    "phone": tutor_profile.phone,
                    "telegram": tutor_profile.telegram,
                    "schedule": tutor_profile.schedule,
                    "status": status_for(relation),
                    "cooldown_until": relation.declined_until,
                }
            )
        if kind == "students":
            return {"students": students_list, "tutors": []}
        if kind == "tutors":
            return {"students": [], "tutors": tutors_list}
        return {"students": students_list, "tutors": tutors_list}

    return {"students": [], "tutors": []}


@router.post("/request")
async def send_tutor_request(
    payload: TutorRequestCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role not in {UserRole.student, UserRole.tutor}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only students or tutors")

    result = await db.execute(select(TutorProfile).where(TutorProfile.id == payload.tutor_id))
    tutor_profile = result.scalar_one_or_none()
    if not tutor_profile:
        raise HTTPException(status_code=404, detail="Tutor not found")

    existing = await db.execute(
        select(TutorStudent).where(
            TutorStudent.tutor_id == payload.tutor_id,
            TutorStudent.student_id == current_user.id,
        )
    )
    relation = existing.scalar_one_or_none()
    now = datetime.now(tz=ZoneInfo("Asia/Bishkek"))
    if relation:
        if relation.confirmed:
            return {"status": "already_confirmed"}
        if relation.declined_until and relation.declined_until > now:
            remaining = int((relation.declined_until - now).total_seconds())
            return {
                "status": "cooldown",
                "cooldown_until": relation.declined_until.isoformat(),
                "cooldown_seconds": remaining,
            }
        if relation.declined_until and relation.declined_until <= now:
            relation.declined_until = None
        else:
            return {"status": "already_requested"}

    if not relation:
        relation = TutorStudent(tutor_id=payload.tutor_id, student_id=current_user.id, confirmed=False)
        db.add(relation)

    message = (
        f"New request from {current_user.first_name} {current_user.last_name} "
        f"for subject {tutor_profile.subject or ''}."
    )
    db.add(
        Notification(
            user_id=tutor_profile.user_id,
            message=message,
            notification_type="tutor_request",
            payload={"student_id": current_user.id, "tutor_id": tutor_profile.id},
        )
    )
    await db.commit()
    return {"status": "ok"}


@router.post("/confirm")
async def confirm_request(
    payload: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    student_id = payload.get("student_id")
    tutor_id = payload.get("tutor_id")
    notification_id = payload.get("notification_id")
    if not student_id or not tutor_id:
        raise HTTPException(status_code=400, detail="student_id and tutor_id required")

    resolved_tutor_profile_id = await _resolve_tutor_profile_id(db, int(tutor_id))
    if resolved_tutor_profile_id is None:
        raise HTTPException(status_code=404, detail="Tutor profile not found")

    owner_result = await db.execute(
        select(TutorProfile.user_id).where(TutorProfile.id == resolved_tutor_profile_id)
    )
    tutor_owner_id = owner_result.scalar_one_or_none()
    if tutor_owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your request")

    student_result = await db.execute(
        select(User.first_name, User.last_name).where(User.id == int(student_id))
    )
    student_row = student_result.first()
    student_first_name = (student_row[0] if student_row else "") or ""
    student_last_name = (student_row[1] if student_row else "") or ""

    relation_result = await db.execute(
        select(TutorStudent).where(
            TutorStudent.tutor_id == resolved_tutor_profile_id,
            TutorStudent.student_id == int(student_id),
        )
    )
    relation = relation_result.scalar_one_or_none()
    already_confirmed = bool(relation and relation.confirmed)
    if not relation:
        relation = TutorStudent(
            tutor_id=resolved_tutor_profile_id,
            student_id=int(student_id),
            confirmed=True,
        )
        db.add(relation)
        await db.flush()
    else:
        relation.confirmed = True
    relation.declined_until = None

    updated_notification: Notification | None = None
    notification_response: dict | None = None
    if notification_id:
        notification_result = await db.execute(
            select(Notification).where(
                Notification.id == notification_id,
                Notification.user_id == current_user.id,
            )
        )
        notification = notification_result.scalar_one_or_none()
        if not notification:
            raise HTTPException(status_code=404, detail="Notification not found")
        notification.is_read = True
        notification.notification_type = "tutor_request_resolved"
        notification.message = (
            f"Вы приняли тьюти: {student_first_name} {student_last_name}."
        )
        notification.payload = {
            "student_id": relation.student_id,
            "tutor_id": relation.tutor_id,
            "action": "confirm",
        }
        updated_notification = notification
        # Build response payload BEFORE commit to avoid attribute expiration triggering IO (MissingGreenlet)
        notification_response = {
            "id": notification.id,
            "message": notification.message,
            "notification_type": notification.notification_type,
            "is_read": notification.is_read,
            "created_at": notification.created_at,
            "payload": notification.payload,
        }

    if not already_confirmed:
        db.add(
            Notification(
                user_id=relation.student_id,
                message=f"Ваш запрос принят тьютором {current_user.first_name} {current_user.last_name}.",
                notification_type="tutor_response",
            )
        )
    await db.commit()
    return {
        "success": True,
        "notification": notification_response,
    }


@router.post("/decline")
async def decline_request(
    payload: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    student_id = payload.get("student_id")
    tutor_id = payload.get("tutor_id")
    notification_id = payload.get("notification_id")
    raw_reason = payload.get("reason")
    if not student_id or not tutor_id:
        raise HTTPException(status_code=400, detail="student_id and tutor_id required")
    reason = str(raw_reason or "").strip()
    if not reason:
        raise HTTPException(status_code=400, detail="reason required")

    resolved_tutor_profile_id = await _resolve_tutor_profile_id(db, int(tutor_id))
    if resolved_tutor_profile_id is None:
        raise HTTPException(status_code=404, detail="Tutor profile not found")

    owner_result = await db.execute(
        select(TutorProfile.user_id).where(TutorProfile.id == resolved_tutor_profile_id)
    )
    tutor_owner_id = owner_result.scalar_one_or_none()
    if tutor_owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your request")

    student_result = await db.execute(
        select(User.first_name, User.last_name).where(User.id == int(student_id))
    )
    student_row = student_result.first()
    student_first_name = (student_row[0] if student_row else "") or ""
    student_last_name = (student_row[1] if student_row else "") or ""

    relation_result = await db.execute(
        select(TutorStudent).where(
            TutorStudent.tutor_id == resolved_tutor_profile_id,
            TutorStudent.student_id == int(student_id),
        )
    )
    relation = relation_result.scalar_one_or_none()
    decline_until = datetime.now(tz=ZoneInfo("Asia/Bishkek")) + timedelta(days=1)
    if not relation:
        relation = TutorStudent(
            tutor_id=resolved_tutor_profile_id,
            student_id=int(student_id),
            confirmed=False,
        )
        db.add(relation)
    relation.confirmed = False
    relation.declined_until = decline_until

    updated_notification: Notification | None = None
    notification_response: dict | None = None
    if notification_id:
        notification_result = await db.execute(
            select(Notification).where(
                Notification.id == notification_id,
                Notification.user_id == current_user.id,
            )
        )
        notification = notification_result.scalar_one_or_none()
        if not notification:
            raise HTTPException(status_code=404, detail="Notification not found")
        notification.is_read = True
        notification.notification_type = "tutor_request_resolved"
        notification.message = (
            f"Вы отклонили тьюти: {student_first_name} {student_last_name}. "
            f"Причина: {reason}"
        )
        notification.payload = {
            "student_id": int(student_id),
            "tutor_id": resolved_tutor_profile_id,
            "action": "decline",
            "reason": reason,
        }
        updated_notification = notification
        # Build response payload BEFORE commit to avoid attribute expiration triggering IO (MissingGreenlet)
        notification_response = {
            "id": notification.id,
            "message": notification.message,
            "notification_type": notification.notification_type,
            "is_read": notification.is_read,
            "created_at": notification.created_at,
            "payload": notification.payload,
        }

    db.add(
        Notification(
            user_id=int(student_id),
            message=f"Ваш запрос отклонён. Причина: {reason}",
            notification_type="tutor_response",
            payload={"reason": reason},
        )
    )
    await db.commit()
    return {
        "success": True,
        "notification": notification_response,
    }


@router.get("/{tutor_id}/schedule")
async def get_tutor_schedule(tutor_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(TutorProfile).where(TutorProfile.id == tutor_id))
    tutor = result.scalar_one_or_none()
    if not tutor:
        raise HTTPException(status_code=404, detail="Tutor not found")
    return tutor.schedule or {}


@router.post("/attendance")
async def submit_attendance(
    payload: AttendanceSubmit,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(TutorStudent).where(
            TutorStudent.tutor_id == payload.tutor_id,
            TutorStudent.student_id == current_user.id,
            TutorStudent.confirmed == True,  # noqa: E712
        )
    )
    relation = result.scalar_one_or_none()
    if not relation:
        raise HTTPException(status_code=404, detail="Request not found")

    relation.attended = True
    relation.attendance_time = datetime.now(tz=ZoneInfo("Asia/Bishkek"))
    await db.commit()
    return {"success": True}


@router.post("/lesson-details")
async def save_lesson_details(
    payload: LessonDetailsUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == payload.tutor_id))
    tutor = result.scalar_one_or_none()
    if not tutor:
        raise HTTPException(status_code=404, detail="Tutor not found")

    tutor.last_lesson_topic = payload.topic
    tutor.last_lesson_location = payload.location
    await db.commit()
    return {"success": True}


@router.post("/lesson-attendance")
async def save_lesson_attendance(
    payload: LessonAttendanceUpdate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == current_user.id))
    tutor = result.scalar_one_or_none()
    if not tutor:
        raise HTTPException(status_code=404, detail="Tutor not found")

    now = time_now(request)
    schedule_entries = _normalize_schedule_entries(tutor.schedule or {})
    active_window: tuple[datetime, datetime] | None = None
    active_entry = None
    for entry in schedule_entries:
        window = _lesson_window(now, entry)
        if not window:
            continue
        start, end = window
        window_end = end + timedelta(hours=1)
        if start <= now <= window_end:
            active_window = (start, window_end)
            active_entry = entry
            break

    if not active_window:
        raise HTTPException(status_code=400, detail="Окно отметок закрыто")

    tutor.last_lesson_topic = payload.topic.strip()
    tutor.last_lesson_location = payload.location.strip() or (active_entry or {}).get("location") or ""

    relations_result = await db.execute(
        select(TutorStudent).where(
            TutorStudent.tutor_id == tutor.id,
            TutorStudent.confirmed == True,  # noqa: E712
        )
    )
    relations = relations_result.scalars().all()
    present_ids = {int(student_id) for student_id in payload.student_ids}
    for relation in relations:
        relation.attended = relation.student_id in present_ids
        relation.attendance_time = now

    db.add(tutor)
    await db.commit()
    return {"success": True}


@router.get("/my-lessons")
async def get_my_lessons(
    tutor_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == tutor_id))
    tutor = result.scalar_one_or_none()
    if not tutor:
        raise HTTPException(status_code=404, detail="Tutor not found")

    relations_result = await db.execute(
        select(TutorStudent)
        .options(selectinload(TutorStudent.student))
        .where(TutorStudent.tutor_id == tutor.id, TutorStudent.confirmed == True)  # noqa: E712
    )
    relations = relations_result.scalars().all()

    now = time_now(request)
    schedule_entries = _normalize_schedule_entries(tutor.schedule or {})

    active_entry = None
    active_window: tuple[datetime, datetime] | None = None
    last_finished_entry = None
    last_finished_window: tuple[datetime, datetime] | None = None
    for entry in schedule_entries:
        window = _lesson_window(now, entry)
        if not window:
            continue
        start, end = window
        window_end = end + timedelta(hours=1)
        if start <= now <= window_end:
            active_entry = entry
            active_window = (start, window_end)
            break
        if window_end < now:
            if not last_finished_window or start > last_finished_window[0]:
                last_finished_entry = entry
                last_finished_window = (start, window_end)

    if active_window and tutor.excel_generated:
        tutor.excel_generated = False
        db.add(tutor)
        await db.commit()

    if not active_window and last_finished_window and not tutor.excel_generated:
        window_start, window_end = last_finished_window
        any_marked = any(
            rel.attendance_time
            and window_start <= rel.attendance_time <= window_end
            and rel.attended
            for rel in relations
        )
        if not any_marked:
            for rel in relations:
                rel.attended = True
                rel.attendance_time = window_end
        records = []
        for rel in relations:
            student = rel.student
            records.append(
                {
                    "first_name": student.first_name or "",
                    "last_name": student.last_name or "",
                    "group": student.group or "",
                    "attended": rel.attended,
                }
            )
        lesson_time = ""
        if last_finished_entry and last_finished_entry.get("pair") in LESSON_TIMES:
            lesson_time = LESSON_TIMES[last_finished_entry.get("pair")]
        file_name = generate_excel_for_lesson(
            tutor=tutor,
            lesson_date=window_start,
            lesson_time=lesson_time,
            records=records,
        )
        tutor.excel_generated = True
        tutor.last_excel_path = file_name
        db.add(tutor)
        await db.commit()

    def is_marked_in_window(relation: TutorStudent, window: tuple[datetime, datetime] | None) -> bool:
        if not window or not relation.attendance_time:
            return False
        return window[0] <= relation.attendance_time <= window[1] and relation.attended

    students = []
    for relation in relations:
        student = relation.student
        students.append(
            {
                "id": student.id,
                "name": f"{student.first_name} {student.last_name}",
                "group": student.group or "",
                "attended": is_marked_in_window(relation, active_window),
                "status": "confirmed",
                "cooldown_until": relation.declined_until,
            }
        )

    return {
        "students": students,
        "show_fields": active_window is not None,
        "topic": tutor.last_lesson_topic or "",
        "location": (active_entry or {}).get("location") or tutor.last_lesson_location or "",
        "active_lesson": {
            "day": active_entry.get("day") if active_entry else "",
            "time": LESSON_TIMES.get(active_entry.get("pair", ""), "") if active_entry else "",
            "location": (active_entry or {}).get("location") or "",
        }
        if active_entry
        else None,
    }


@router.get("/schedule-with-attendance", response_model=TutorScheduleResponse)
async def get_schedule_with_attendance(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(TutorStudent)
        .where(TutorStudent.student_id == current_user.id)
        .where(TutorStudent.confirmed == True)  # noqa: E712
    )
    relations = result.scalars().all()
    if not relations:
        return {"schedule": []}

    now = datetime.now(tz=ZoneInfo("Asia/Bishkek"))
    current_day_index = now.weekday()

    schedule_list = []
    for relation in relations:
        tutor = relation.tutor
        schedule = tutor.schedule or {}
        formatted = []
        for entry in _normalize_schedule_entries(schedule):
            pair_num = entry.get("pair", "")
            day_name = entry.get("day", "")
            if pair_num not in LESSON_TIMES:
                continue
            window = _lesson_window(now, entry)
            if window:
                start_time, end_time = window
                allowed = start_time <= now <= (end_time + timedelta(hours=1))
            else:
                allowed = False
            formatted.append(
                {
                    "day": day_name,
                    "time": LESSON_TIMES[pair_num],
                    "can_mark": allowed,
                }
            )
        schedule_list.append(
            {
                "tutor_id": tutor.id,
                "tutor_name": f"{tutor.user.first_name} {tutor.user.last_name}",
                "subject": tutor.subject,
                "schedule": formatted,
                "already_marked": relation.attended,
            }
        )

    return {"schedule": schedule_list}


def generate_excel_for_lesson(
    *,
    tutor: TutorProfile,
    lesson_date: datetime,
    lesson_time: str,
    records: list[dict],
) -> str:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    from openpyxl.formula.translate import Translator

    safe_first = (tutor.user.first_name or "").strip().replace(" ", "_")
    safe_last = (tutor.user.last_name or "").strip().replace(" ", "_")
    safe_first = "".join(ch for ch in safe_first if ch.isalnum() or ch in "_-")
    safe_last = "".join(ch for ch in safe_last if ch.isalnum() or ch in "_-")
    file_name = f"report_{safe_first}{safe_last}_{tutor.user_id}.xlsx"
    file_path = LESSON_DIR / file_name

    formula_seed: dict[int, str] = {}
    if file_path.exists():
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        template_path = (
            Path(settings.report_template_path)
            if settings.report_template_path
            else (Path(__file__).resolve().parents[5] / "Report.xlsx")
        )
        if template_path.exists():
            workbook = load_workbook(template_path)
            sheet = workbook.active
            if sheet.max_row >= 6:
                for cell in sheet[6]:
                    if cell.data_type == "f" and cell.value:
                        formula_seed[cell.column] = cell.value
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"
            sheet.append(
                [
                    "Дата",
                    "Время",
                    "Предмет",
                    "Тема",
                    "Кабинет",
                    "Фамилия",
                    "Имя",
                    "Группа",
                    "Был",
                ]
            )
            for cell in sheet[1]:
                cell.font = Font(bold=True)

        if sheet.max_row >= 6:
            sheet.delete_rows(6, sheet.max_row - 5)

        tutor_name = f"{tutor.user.first_name or ''} {tutor.user.last_name or ''}".strip()
        if sheet["B2"].value:
            sheet["B2"] = f"Тьютор: {tutor_name}"
        if sheet["B3"].value:
            sheet["B3"] = f"Предмет: {tutor.subject or ''}"

    formula_map: dict[int, str] = {}
    if sheet.max_row >= 6:
        for cell in sheet[6]:
            if cell.data_type == "f" and cell.value:
                formula_map[cell.column] = cell.value
    if not formula_map and formula_seed:
        formula_map = formula_seed

    header_row = 5
    date_column = None
    for col in range(4, sheet.max_column + 1):
        value = sheet.cell(row=header_row, column=col).value
        if value is None or (isinstance(value, str) and "дд.мм" in value):
            if date_column is None:
                date_column = col
        elif isinstance(value, datetime):
            if value.date() == lesson_date.date():
                date_column = col
                break
    if date_column is None:
        date_column = sheet.max_column + 1
    sheet.cell(row=header_row, column=date_column).value = lesson_date.date()

    for record in records:
        name_value = f"{record.get('last_name', '').strip()} {record.get('first_name', '').strip()}".strip()
        group_value = record.get("group", "").strip()
        target_row = None
        for row in range(6, sheet.max_row + 1):
            if (
                str(sheet.cell(row=row, column=2).value or "").strip() == name_value
                and str(sheet.cell(row=row, column=3).value or "").strip() == group_value
            ):
                target_row = row
                break
        if target_row is None:
            target_row = max(sheet.max_row + 1, 6)
            sheet.cell(row=target_row, column=1).value = '=$B$2&"/"&$B$3'
            sheet.cell(row=target_row, column=2).value = name_value
            sheet.cell(row=target_row, column=3).value = group_value
            for col, formula in formula_map.items():
                try:
                    sheet.cell(row=target_row, column=col).value = Translator(
                        formula, origin=f"{sheet.cell(row=6, column=col).coordinate}"
                    ).translate_formula(
                        f"{sheet.cell(row=target_row, column=col).coordinate}"
                    )
                except Exception:
                    sheet.cell(row=target_row, column=col).value = formula

        sheet.cell(row=target_row, column=date_column).value = 1 if record.get("attended") else 0

    workbook.save(file_path)
    return str(file_path)


@router.get("/lesson-file/{tutor_id}")
async def get_lesson_file(
    tutor_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == tutor_id))
    tutor = result.scalar_one_or_none()
    if not tutor or not tutor.last_excel_path:
        raise HTTPException(status_code=404, detail="Lesson file not found")

    return FileResponse(
        tutor.last_excel_path,
        filename=Path(tutor.last_excel_path).name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
