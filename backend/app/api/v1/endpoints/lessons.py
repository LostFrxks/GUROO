from __future__ import annotations

from datetime import date, datetime, time, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import FileResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import get_current_user
from app.core.config import settings
from app.core.time import now as now_time
from app.db.session import get_db
from app.models.lesson import AttendanceStatus, EnrollmentStatus, ExamScore, Lesson, LessonEnrollment
from app.models.tutor import Notification, TutorProfile
from app.models.user import User, UserRole
from app.schemas.lesson import (
    AttendanceUpdate,
    LessonDetailsUpdate,
    LessonOut,
    LessonRequestCreate,
    LessonRequestResponse,
    StudentLessonItem,
    StudentRsvp,
    TutorDecision,
    TutorLessonItem,
)

router = APIRouter()

TIMEZONE = ZoneInfo("Asia/Bishkek")
BOOKING_OPEN_TIME = time(8, 0)
LESSON_DURATION_MINUTES = 80
ATTENDANCE_WINDOW_BEFORE_END_MINUTES = 10
ATTENDANCE_WINDOW_AFTER_END_HOURS = 1

LESSON_TIMES = {
    "1": "08:00-09:20",
    "2": "09:30-10:50",
    "3": "11:40-13:00",
    "4": "13:10-14:30",
    "5": "14:40-16:00",
    "6": "16:00-17:20",
    "7": "17:30-19:00",
}

DAY_RU = {
    "Monday": "Понедельник",
    "Tuesday": "Вторник",
    "Wednesday": "Среда",
    "Thursday": "Четверг",
    "Friday": "Пятница",
    "Saturday": "Суббота",
    "Sunday": "Воскресенье",
}
DAY_EN = {value: key for key, value in DAY_RU.items()}

ROOM_OPTIONS = {"Балкон"} | {f"C{i}" for i in range(1, 12)} | {f"LAB{i}" for i in range(1, 4)}

LESSON_DIR = Path(settings.media_root) / "lessons"
LESSON_DIR.mkdir(parents=True, exist_ok=True)

REPORT_TEMPLATE_NAME = "Отчетный_материал_по_ТЬЮТОРАМ_2023_2025_УГ.xlsx"


def _get_report_template_path() -> Path:
    if settings.report_template_path:
        candidate = Path(settings.report_template_path)
        if not candidate.is_absolute():
            candidate = (Path.cwd() / candidate).resolve()
        return candidate
    for parent in Path(__file__).resolve().parents:
        candidate = parent / REPORT_TEMPLATE_NAME
        if candidate.exists():
            return candidate
    return (Path.cwd() / REPORT_TEMPLATE_NAME).resolve()


def _lesson_start_at(lesson_date: date, slot_id: str) -> datetime:
    if slot_id not in LESSON_TIMES:
        raise HTTPException(status_code=400, detail="Invalid lesson slot")
    start_str, _ = LESSON_TIMES[slot_id].split("-")
    start_hour, start_min = map(int, start_str.split(":"))
    return datetime.combine(lesson_date, time(start_hour, start_min), tzinfo=TIMEZONE)


def _attendance_window(lesson_date: date, slot_id: str) -> tuple[datetime, datetime]:
    start_at = _lesson_start_at(lesson_date, slot_id)
    end_at = start_at + timedelta(minutes=LESSON_DURATION_MINUTES)
    window_start = end_at - timedelta(minutes=ATTENDANCE_WINDOW_BEFORE_END_MINUTES)
    window_end = end_at + timedelta(hours=ATTENDANCE_WINDOW_AFTER_END_HOURS)
    return window_start, window_end


def _schedule_has_slot(schedule: dict, lesson_date: date, slot_id: str) -> bool:
    weekday_en = lesson_date.strftime("%A")
    weekday_ru = DAY_RU.get(weekday_en, weekday_en)
    return slot_id in (schedule.get(weekday_en, []) or []) or slot_id in (schedule.get(weekday_ru, []) or [])


def _ensure_booking_window(now: datetime, lesson_date: date) -> None:
    tomorrow = now.date() + timedelta(days=1)
    if lesson_date != tomorrow:
        raise HTTPException(status_code=400, detail="Booking is allowed only for tomorrow")
    if now.time() < BOOKING_OPEN_TIME:
        raise HTTPException(status_code=400, detail="Booking opens at 08:00 the day before")


def _ensure_rsvp_window(now: datetime, lesson_date: date) -> None:
    if now.date() != lesson_date - timedelta(days=1):
        raise HTTPException(status_code=400, detail="RSVP is allowed only the day before the lesson")
    if now.time() < BOOKING_OPEN_TIME:
        raise HTTPException(status_code=400, detail="RSVP opens at 08:00")


def _ensure_cancel_window(now: datetime, lesson_date: date, slot_id: str) -> None:
    start_at = _lesson_start_at(lesson_date, slot_id)
    if now > start_at - timedelta(hours=24):
        raise HTTPException(status_code=400, detail="Cancelation is closed 24 hours before the lesson")


async def _weekly_enrollment_count(
    db: AsyncSession, *, tutor_id: int, student_id: int, lesson_date: date
) -> int:
    week_start = lesson_date - timedelta(days=lesson_date.weekday())
    week_end = week_start + timedelta(days=6)
    stmt = (
        select(func.count(LessonEnrollment.id))
        .join(Lesson)
        .where(Lesson.tutor_id == tutor_id)
        .where(Lesson.lesson_date >= week_start, Lesson.lesson_date <= week_end)
        .where(LessonEnrollment.student_id == student_id)
        .where(
            LessonEnrollment.status.notin_(
                [EnrollmentStatus.tutor_declined, EnrollmentStatus.student_declined, EnrollmentStatus.canceled]
            )
        )
    )
    result = await db.execute(stmt)
    return int(result.scalar_one() or 0)


async def _auto_mark_attendance(db: AsyncSession, lesson: Lesson) -> None:
    if lesson.attendance_finalized:
        return
    any_marked = any(
        enrollment.attendance_status != AttendanceStatus.unmarked
        for enrollment in lesson.enrollments
    )
    if any_marked:
        lesson.attendance_finalized = True
        db.add(lesson)
        await db.commit()
        return
    for enrollment in lesson.enrollments:
        if enrollment.status == EnrollmentStatus.student_confirmed:
            enrollment.attendance_status = AttendanceStatus.present
            enrollment.attendance_marked_at = datetime.now(tz=TIMEZONE)
    lesson.attendance_finalized = True
    db.add(lesson)
    await db.commit()


async def _auto_close_rsvp(db: AsyncSession, lesson: Lesson, now: datetime) -> None:
    if now.date() < lesson.lesson_date:
        return
    updated = False
    for enrollment in lesson.enrollments:
        if enrollment.status == EnrollmentStatus.tutor_accepted:
            enrollment.status = EnrollmentStatus.student_declined
            updated = True
            db.add(enrollment)
    if updated:
        await db.commit()


@router.post("/request", response_model=LessonRequestResponse)
async def request_lesson(
    payload: LessonRequestCreate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.student:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Students only")

    now = now_time(request)
    _ensure_booking_window(now, payload.lesson_date)

    tutor_result = await db.execute(select(TutorProfile).where(TutorProfile.id == payload.tutor_id))
    tutor = tutor_result.scalar_one_or_none()
    if not tutor:
        raise HTTPException(status_code=404, detail="Tutor not found")

    if not _schedule_has_slot(tutor.schedule or {}, payload.lesson_date, payload.slot_id):
        raise HTTPException(status_code=400, detail="Tutor is not available for this slot")

    week_count = await _weekly_enrollment_count(
        db,
        tutor_id=tutor.id,
        student_id=current_user.id,
        lesson_date=payload.lesson_date,
    )
    if week_count >= 2:
        raise HTTPException(status_code=400, detail="Weekly limit reached (2 lessons)")

    lesson_result = await db.execute(
        select(Lesson).where(
            Lesson.tutor_id == tutor.id,
            Lesson.lesson_date == payload.lesson_date,
            Lesson.slot_id == payload.slot_id,
        )
    )
    lesson = lesson_result.scalar_one_or_none()
    if not lesson:
        lesson = Lesson(tutor_id=tutor.id, lesson_date=payload.lesson_date, slot_id=payload.slot_id)
        db.add(lesson)
        await db.flush()

    existing = await db.execute(
        select(LessonEnrollment).where(
            LessonEnrollment.lesson_id == lesson.id,
            LessonEnrollment.student_id == current_user.id,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail="Request already exists")

    enrollment = LessonEnrollment(lesson_id=lesson.id, student_id=current_user.id)
    db.add(enrollment)
    await db.flush()

    message = (
        f"📩 Новый запрос на занятие от {current_user.first_name} {current_user.last_name} "
        f"по предмету «{tutor.subject or ''}» на {payload.lesson_date} ({LESSON_TIMES[payload.slot_id]})."
    )
    db.add(
        Notification(
            user_id=tutor.user_id,
            message=message,
            notification_type="lesson_request",
            payload={
                "enrollment_id": enrollment.id,
                "student_id": current_user.id,
                "lesson_date": str(payload.lesson_date),
                "slot_id": payload.slot_id,
            },
        )
    )
    await db.commit()
    await db.refresh(enrollment)

    payment_notice = week_count == 1 and not current_user.hide_payment_notice
    return LessonRequestResponse(
        enrollment_id=enrollment.id,
        status=enrollment.status,
        payment_notice=payment_notice,
    )


@router.post("/{enrollment_id}/accept")
async def accept_request(
    enrollment_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    result = await db.execute(
        select(LessonEnrollment)
        .where(LessonEnrollment.id == enrollment_id)
        .options(
            selectinload(LessonEnrollment.lesson)
            .selectinload(Lesson.tutor)
            .selectinload(TutorProfile.user)
        )
    )
    enrollment = result.scalar_one_or_none()
    if not enrollment:
        raise HTTPException(status_code=404, detail="Request not found")

    if enrollment.lesson.tutor.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your request")

    if enrollment.status != EnrollmentStatus.requested:
        raise HTTPException(status_code=400, detail="Request already processed")

    enrollment.status = EnrollmentStatus.tutor_accepted
    db.add(enrollment)
    db.add(
        Notification(
            user_id=enrollment.student_id,
            message=(
                "✅ Тьютор подтвердил вашу запись. Подтвердите посещение за день до урока в 08:00."
            ),
            notification_type="lesson_accepted",
            payload={"enrollment_id": enrollment.id},
        )
    )
    await db.commit()
    return {"success": True}


@router.post("/{enrollment_id}/decline")
async def decline_request(
    enrollment_id: int,
    payload: TutorDecision,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    if not payload.reason:
        raise HTTPException(status_code=400, detail="Decline reason required")

    result = await db.execute(
        select(LessonEnrollment)
        .where(LessonEnrollment.id == enrollment_id)
        .options(
            selectinload(LessonEnrollment.lesson)
            .selectinload(Lesson.tutor)
            .selectinload(TutorProfile.user)
        )
    )
    enrollment = result.scalar_one_or_none()
    if not enrollment:
        raise HTTPException(status_code=404, detail="Request not found")

    if enrollment.lesson.tutor.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your request")

    enrollment.status = EnrollmentStatus.tutor_declined
    enrollment.decline_reason = payload.reason
    db.add(enrollment)
    db.add(
        Notification(
            user_id=enrollment.student_id,
            message=f"❌ Тьютор отклонил вашу запись. Причина: {payload.reason}",
            notification_type="lesson_declined",
            payload={"enrollment_id": enrollment.id},
        )
    )
    await db.commit()
    return {"success": True}


@router.post("/{enrollment_id}/rsvp")
async def rsvp(
    enrollment_id: int,
    payload: StudentRsvp,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(LessonEnrollment)
        .where(LessonEnrollment.id == enrollment_id)
        .options(
            selectinload(LessonEnrollment.lesson)
            .selectinload(Lesson.tutor)
            .selectinload(TutorProfile.user)
        )
    )
    enrollment = result.scalar_one_or_none()
    if not enrollment:
        raise HTTPException(status_code=404, detail="Request not found")
    if enrollment.student_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your request")
    if enrollment.status != EnrollmentStatus.tutor_accepted:
        raise HTTPException(status_code=400, detail="Request is not accepted yet")

    now = now_time(request)
    _ensure_rsvp_window(now, enrollment.lesson.lesson_date)

    if payload.will_attend:
        enrollment.status = EnrollmentStatus.student_confirmed
        message = "✅ Студент подтвердил посещение."
        notification_type = "lesson_student_confirmed"
    else:
        enrollment.status = EnrollmentStatus.student_declined
        message = "❌ Студент отказался от посещения."
        notification_type = "lesson_student_declined"

    db.add(enrollment)
    db.add(
        Notification(
            user_id=enrollment.lesson.tutor.user_id,
            message=message,
            notification_type=notification_type,
            payload={"enrollment_id": enrollment.id},
        )
    )
    await db.commit()
    return {"success": True}


@router.post("/{enrollment_id}/cancel")
async def cancel_request(
    enrollment_id: int,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(LessonEnrollment)
        .where(LessonEnrollment.id == enrollment_id)
        .options(
            selectinload(LessonEnrollment.lesson)
            .selectinload(Lesson.tutor)
            .selectinload(TutorProfile.user)
        )
    )
    enrollment = result.scalar_one_or_none()
    if not enrollment:
        raise HTTPException(status_code=404, detail="Request not found")
    if enrollment.student_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your request")

    now = now_time(request)
    _ensure_cancel_window(now, enrollment.lesson.lesson_date, enrollment.lesson.slot_id)

    if enrollment.status in {EnrollmentStatus.tutor_declined, EnrollmentStatus.canceled}:
        raise HTTPException(status_code=400, detail="Request already closed")

    enrollment.status = EnrollmentStatus.canceled
    db.add(enrollment)
    db.add(
        Notification(
            user_id=enrollment.lesson.tutor.user_id,
            message="ℹ️ Студент отменил запись.",
            notification_type="lesson_canceled",
            payload={"enrollment_id": enrollment.id},
        )
    )
    await db.commit()
    return {"success": True}


@router.get("/student", response_model=list[StudentLessonItem])
async def list_student_lessons(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.student:
        raise HTTPException(status_code=403, detail="Students only")

    result = await db.execute(
        select(LessonEnrollment)
        .join(Lesson)
        .where(LessonEnrollment.student_id == current_user.id)
        .order_by(Lesson.lesson_date.desc())
        .options(
            selectinload(LessonEnrollment.lesson)
            .selectinload(Lesson.tutor)
            .selectinload(TutorProfile.user)
        )
    )
    items = []
    now = datetime.now(tz=TIMEZONE)
    for enrollment in result.scalars().all():
        await _auto_close_rsvp(db, enrollment.lesson, now)
        tutor = enrollment.lesson.tutor
        tutor_user = tutor.user
        items.append(
            StudentLessonItem(
                id=enrollment.id,
                tutor_id=tutor.id,
                tutor_name=f"{tutor_user.first_name} {tutor_user.last_name}",
                subject=tutor.subject,
                lesson_date=enrollment.lesson.lesson_date,
                slot_id=enrollment.lesson.slot_id,
                status=enrollment.status,
                decline_reason=enrollment.decline_reason,
                attendance_status=enrollment.attendance_status,
                topic=enrollment.lesson.topic,
                location=enrollment.lesson.location,
            )
        )
    return items


@router.get("/tutor", response_model=list[TutorLessonItem])
async def list_tutor_lessons(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    result = await db.execute(
        select(Lesson)
        .join(TutorProfile)
        .where(TutorProfile.user_id == current_user.id)
        .order_by(Lesson.lesson_date.desc())
        .options(selectinload(Lesson.enrollments).selectinload(LessonEnrollment.student))
    )
    lessons = result.scalars().unique().all()

    now = now_time(request)
    for lesson in lessons:
        await _auto_close_rsvp(db, lesson, now)
        window_start, window_end = _attendance_window(lesson.lesson_date, lesson.slot_id)
        if now > window_end:
            await _auto_mark_attendance(db, lesson)

    return lessons


@router.get("/{lesson_id}", response_model=LessonOut)
async def get_lesson(
    lesson_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(Lesson)
        .where(Lesson.id == lesson_id)
        .options(selectinload(Lesson.enrollments).selectinload(LessonEnrollment.student))
    )
    lesson = result.scalar_one_or_none()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    if current_user.role == UserRole.tutor and lesson.tutor.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your lesson")
    if current_user.role == UserRole.student:
        student_ids = {e.student_id for e in lesson.enrollments}
        if current_user.id not in student_ids:
            raise HTTPException(status_code=403, detail="Not your lesson")
    return lesson


@router.post("/{lesson_id}/details")
async def update_lesson_details(
    lesson_id: int,
    payload: LessonDetailsUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if payload.location not in ROOM_OPTIONS:
        raise HTTPException(status_code=400, detail="Invalid location")

    result = await db.execute(
        select(Lesson)
        .where(Lesson.id == lesson_id)
        .options(selectinload(Lesson.enrollments))
    )
    lesson = result.scalar_one_or_none()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    if lesson.tutor.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your lesson")

    lesson.topic = payload.topic
    lesson.location = payload.location
    db.add(lesson)
    await db.commit()
    return {"success": True}


@router.post("/{lesson_id}/attendance")
async def mark_attendance(
    lesson_id: int,
    payload: AttendanceUpdate,
    request: Request,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Lesson).where(Lesson.id == lesson_id))
    lesson = result.scalar_one_or_none()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    if lesson.tutor.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not your lesson")

    now = now_time(request)
    window_start, window_end = _attendance_window(lesson.lesson_date, lesson.slot_id)
    if not (window_start <= now <= window_end):
        raise HTTPException(status_code=403, detail="Attendance window is closed")

    present_ids = set(payload.present_student_ids)
    for enrollment in lesson.enrollments:
        if enrollment.status != EnrollmentStatus.student_confirmed:
            continue
        if enrollment.student_id in present_ids:
            enrollment.attendance_status = AttendanceStatus.present
        else:
            enrollment.attendance_status = AttendanceStatus.absent
        enrollment.attendance_marked_at = now
        db.add(enrollment)

    lesson.attendance_finalized = True
    db.add(lesson)
    await db.commit()
    return {"success": True}


@router.post("/scores")
async def upsert_scores(
    payload: dict,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    tutor_id = payload.get("tutor_id")
    subject = payload.get("subject")
    midterm1 = payload.get("midterm1")
    midterm2 = payload.get("midterm2")
    final = payload.get("final")

    if not tutor_id:
        raise HTTPException(status_code=400, detail="tutor_id required")

    for value, limit, label in (
        (midterm1, 10, "midterm1"),
        (midterm2, 10, "midterm2"),
        (final, 25, "final"),
    ):
        if value is None:
            continue
        if not isinstance(value, int) or value < 0 or value > limit:
            raise HTTPException(status_code=400, detail=f"{label} must be 0..{limit}")

    result = await db.execute(
        select(ExamScore).where(
            ExamScore.tutor_id == tutor_id,
            ExamScore.student_id == current_user.id,
            ExamScore.subject == subject,
        )
    )
    score = result.scalar_one_or_none()
    if not score:
        score = ExamScore(
            tutor_id=tutor_id,
            student_id=current_user.id,
            subject=subject,
        )
        db.add(score)

    score.midterm1 = midterm1
    score.midterm2 = midterm2
    score.final = final
    db.add(score)
    await db.commit()
    return {"success": True}


@router.get("/scores/tutor")
async def list_scores_for_tutor(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    tutor_result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == current_user.id))
    tutor = tutor_result.scalar_one_or_none()
    if not tutor:
        raise HTTPException(status_code=404, detail="Tutor profile not found")

    scores_result = await db.execute(select(ExamScore).where(ExamScore.tutor_id == tutor.id))
    scores = scores_result.scalars().all()

    data = []
    for score in scores:
        student = score.student
        data.append(
            {
                "student_id": student.id,
                "first_name": student.first_name,
                "last_name": student.last_name,
                "group": student.group,
                "subject": score.subject,
                "midterm1": score.midterm1,
                "midterm2": score.midterm2,
                "final": score.final,
            }
        )
    return {"items": data}


@router.get("/export")
async def export_lessons_report(
    semester: str,
    midterm1_date: date,
    midterm2_date: date,
    final_date: date,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if current_user.role != UserRole.tutor:
        raise HTTPException(status_code=403, detail="Tutor only")

    tutor_result = await db.execute(select(TutorProfile).where(TutorProfile.user_id == current_user.id))
    tutor = tutor_result.scalar_one_or_none()
    if not tutor:
        raise HTTPException(status_code=404, detail="Tutor profile not found")

    template_path = _get_report_template_path()
    if not template_path.exists():
        raise HTTPException(status_code=500, detail="Report template not found")

    from openpyxl import load_workbook

    wb = load_workbook(template_path)
    if semester not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Semester sheet not found")
    ws = wb[semester]

    enrollments_result = await db.execute(
        select(LessonEnrollment)
        .join(Lesson)
        .where(Lesson.tutor_id == tutor.id)
        .where(LessonEnrollment.attendance_status == AttendanceStatus.present)
    )
    enrollments = enrollments_result.scalars().all()

    rows_by_student: dict[int, dict] = {}
    for enrollment in enrollments:
        student = enrollment.student
        lesson_date = enrollment.lesson.lesson_date
        student_row = rows_by_student.setdefault(
            student.id,
            {
                "fio": f"{student.last_name or ''} {student.first_name or ''}".strip(),
                "id": student.id,
                "group": student.group or "",
                "course": "",
                "subject": tutor.subject or "",
                "tutor": f"{tutor.user.last_name or ''} {tutor.user.first_name or ''}".strip(),
                "semester": semester,
                "before_mid1": 0,
                "between_mid": 0,
                "before_final": 0,
                "hours": 0,
            },
        )
        if lesson_date <= midterm1_date:
            student_row["before_mid1"] += 1
        elif lesson_date <= midterm2_date:
            student_row["between_mid"] += 1
        elif lesson_date <= final_date:
            student_row["before_final"] += 1
        student_row["hours"] += 2

    start_row = 2
    for idx, row in enumerate(rows_by_student.values(), start=start_row):
        ws.cell(row=idx, column=1, value=row["fio"])
        ws.cell(row=idx, column=2, value=row["id"])
        ws.cell(row=idx, column=3, value=row["group"])
        ws.cell(row=idx, column=4, value=row["course"])
        ws.cell(row=idx, column=5, value=row["subject"])
        ws.cell(row=idx, column=6, value=row["tutor"])
        ws.cell(row=idx, column=7, value=row["semester"])
        ws.cell(row=idx, column=8, value=row["before_mid1"])
        ws.cell(row=idx, column=9, value=row["between_mid"])
        ws.cell(row=idx, column=10, value=row["before_final"])
        ws.cell(row=idx, column=11, value=row["hours"])

    file_name = f"report_{tutor.user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = LESSON_DIR / file_name
    wb.save(file_path)
    return FileResponse(
        file_path,
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
