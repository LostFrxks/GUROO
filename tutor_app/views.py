import json
import os
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.core.mail import send_mail
from django.views.decorators.csrf import csrf_exempt
from tutor_app.models import CustomUser, TutorStudent, TutorProfile  # ❌ Убрали `Tutor`
from django.middleware.csrf import get_token
import random
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from .models import TutorProfile, TutorInviteToken
from .forms import TutorProfileForm
from datetime import datetime, timedelta 
import pytz
import uuid
import pymysql
from django.db import connection
from django.db import IntegrityError
import socket
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.conf import settings
from datetime import datetime
from .models import Notification
import openpyxl
from openpyxl.styles import Font
from django.http import FileResponse

DB_HOST = "localhost"
DB_USER = "root"
DB_PASSWORD = "root"
DB_NAME = "tutor_project"


lesson_times = {
    "1": "08:00-09:20",
    "2": "09:30-10:50",
    "3": "11:40-13:00",
    "4": "13:10-14:30",
    "5": "14:40-16:00",
    "6": "16:00-17:20",
    "7": "17:30-19:00"
}


def api_home(request):
    return JsonResponse({"message": "API работает!"})

def main_page(request):
    return render(request, "tutor_app/main.html")

def about_page(request):
    return render(request, "tutor_app/about.html")

def index(request):
    return render(request, "tutor_app/tutors.html")  # Страница тьюторов

def get_server_ip():
    """Функция для получения текущего IP сервера."""
    try:
        return socket.gethostbyname(socket.gethostname())
    except socket.gaierror:
        return "127.0.0.1"  # Фоллбэк, если IP не удалось получить

# ✅ Теперь страница тьюторов НИКОГДА не откроется без авторизации!
@login_required(login_url='/login/')
def index(request):
    if not request.user.is_authenticated:
        return redirect("/login/")  # Перенаправляем сразу на логин

    return render(request, 'tutor_app/tutors.html', {"user": request.user})
# ✅ Страница логина (если авторизован, редирект на главную)
def login_page(request):
    if request.user.is_authenticated:
        return redirect("/")
    return render(request, "tutor_app/sign_in.html")

def get_csrf_token(request):
    return JsonResponse({"csrfToken": get_token(request)})

from django.http import JsonResponse
from .models import TutorProfile, TutorStudent
import pytz
from datetime import datetime, timedelta

def get_tutors(request):
    user_id = request.GET.get("user_id")
    
    tutors = TutorProfile.objects.select_related("user").filter(bio__isnull=False).exclude(bio="")    
    
    confirmed_tutors = {
        reg.tutor_id: reg.attended
        for reg in TutorStudent.objects.filter(student_id=user_id, confirmed=True)
    }

    print("📌 Confirmed tutors:", confirmed_tutors)  # 🔍 Проверяем, какие ID он получает
    
    data = []
    for tutor in tutors:
        try:
            relation = TutorStudent.objects.get(tutor=tutor, student_id=user_id)
            if relation.confirmed:
                request_status = "confirmed"
            else:
                request_status = "pending"
        except TutorStudent.DoesNotExist:
            request_status = None

        tutor_data = {
            "id": tutor.id,
            "user_id": tutor.user.id,
            "name": f"{tutor.user.first_name} {tutor.user.last_name}",
            "subject": tutor.subject,
            "bio": tutor.bio,
            "photo": str(tutor.photo) if tutor.photo else "",
            "is_registered": tutor.id in confirmed_tutors,
            "attended": confirmed_tutors.get(tutor.id, False),
            "schedule": tutor.schedule,
            "course": tutor.course,
            "phone": tutor.phone,
            "telegram": tutor.telegram,
            "group": tutor.group,
            "own_course": tutor.own_course,
            "request_status": request_status  # 🔥 новое поле
        }
        data.append(tutor_data)

    return JsonResponse(data, safe=False)




@csrf_exempt
@login_required(login_url='/login/')
def register_tutor(request, tutor_id, student_id):
    print(f"🔍 Получен запрос на запись: tutor_id={tutor_id}, student_id={student_id}")

    tutor = get_object_or_404(TutorProfile, id=tutor_id)
    student = get_object_or_404(CustomUser, id=student_id)

    existing_registration = TutorStudent.objects.filter(tutor=tutor, student=student).first()

    if existing_registration:
        if existing_registration.confirmed:
            print("⚠️ Вы уже записаны!")
            return JsonResponse({"success": False, "error": "Вы уже записаны к этому тьютору!"}, status=400)
        else:
            print("⚠️ Вы уже отправили запрос!")
            return JsonResponse({"success": False, "error": "Вы уже отправили запрос! Ожидайте подтверждения."}, status=400)

    print("✅ Создаем новую запись...")
    record = TutorStudent.objects.create(tutor=tutor, student=student, confirmed=False)
    
    server_ip = get_server_ip()
    confirm_link = f"http://{server_ip}:8000/confirm_registration/{record.id}/"
    print(f"🔗 Ссылка подтверждения: {confirm_link}")

    send_mail(
        "Запись на занятие",
        f"Пользователь {student.email} хочет записаться. Подтвердите: {confirm_link}",
        "admin@site.com",
        [tutor.user.email]
    )

    print(f"📩 Email отправлен на {tutor.user.email}!")  

    return JsonResponse({"success": True, "message": "Запрос отправлен! Ожидайте подтверждения."})


@csrf_exempt
@login_required(login_url='/login/')
def confirm_registration(request, record_id):
    record = get_object_or_404(TutorStudent, id=record_id)
    record.confirmed = True
    record.save()
    print(f"✅ Запись подтверждена: {record.student.email} -> {record.tutor.user.email}")
    
    return JsonResponse({"success": True, "message": "Запись подтверждена!"})
    
from django.contrib.auth import authenticate, login
from django.views.decorators.csrf import csrf_exempt
import json

User = get_user_model()  # ✅ Теперь Django знает, что у нас кастомный пользователь

# ✅ Логин пользователя (добавил обработку GET-запроса)
def user_login(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        if not email or not password:
            return JsonResponse({"error": "Заполните все поля!"}, status=400)

        user = authenticate(request, username=email, password=password)
        if user is not None:
            login(request, user)
            return JsonResponse({"success": True, "redirect_url": "/tutors/"})  # ✅ Перенаправляем на тьюторов
        else:
            return JsonResponse({"error": "Неверная почта или пароль"}, status=400)

    return render(request, "tutor_app/sign_in.html")

def user_logout(request):
    logout(request)  # Разлогиниваем пользователя

    response = JsonResponse({"success": True, "message": "Вы вышли из системы!"})
    
    # ❗️Очищаем кэш, чтобы кнопка "Назад" не работала
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"

    return response

# ✅ Регистрация пользователя (тоже не менял)
@csrf_exempt
def user_register(request):
    if request.method == "POST":
        data = json.loads(request.body)
        username = data.get("username")
        email = data.get("email")
        password = data.get("password")

        if CustomUser.objects.filter(username=username).exists():   
            return JsonResponse({"success": False, "message": "Такой пользователь уже есть!"})

        user = CustomUser.objects.create_user(username=username, email=email, password=password)
        return JsonResponse({"success": True, "message": "Вы успешно зарегистрировались!"})

    return JsonResponse({"success": False, "message": "Метод не поддерживается!"})

@csrf_exempt
@login_required(login_url='/login/')
def mark_attendance(request, tutor_id):
    print(f"📌 Получен запрос на отметку посещения: tutor_id={tutor_id}")

    # Получаем запись посещения по студенту и тьютору
    record = get_object_or_404(TutorStudent, student_id=request.user.id, tutor_id=tutor_id)

    if not record.confirmed:
        print("❌ Ошибка: Студент еще не записан!")
        return JsonResponse({"success": False, "message": "Вы ещё не записаны!"})

    if record.attended:
        print("✅ Студент уже отметился!")
        return JsonResponse({"success": True, "message": "Вы уже отметились!", "attended": True})

    # Отмечаем посещение
    TutorStudent.objects.filter(id=record.id).update(attended=True)

    print(f"✅ Посещение сохранено! Student ID={record.student.id}, Tutor ID={record.tutor.user.id}")

    return JsonResponse({"success": True, "message": "Вы отметили посещение!", "attended": True})

# РЕГИСТРАЦИЯ

# Хранилище кодов (лучше хранить в БД, но пока так)
verification_codes = {}

def index(request):
    return render(request, 'tutor_app/tutors.html')

def register_page(request):
    return render(request, "tutor_app/register.html")

@csrf_exempt
def send_verification_code(request):
    if request.method == "POST":
        data = json.loads(request.body)
        email = data.get("email", "").strip()

        # ✅ Добавляем SELECT только нужных данных (ускоряет запрос)
        exists = CustomUser.objects.filter(email=email).values_list("id", flat=True).exists()
        if exists:
            return JsonResponse({"success": False, "error": "Пользователь с таким email уже зарегистрирован!"})

        # ✅ Дальше процесс без изменений
        verification_code = random.randint(100000, 999999)
        verification_codes[email] = verification_code  

        send_mail(
            "Код подтверждения",
            f"Ваш код: {verification_code}",
            "admin@site.com",
            [email],
            fail_silently=False,
        )

        return JsonResponse({"success": True, "message": "Код отправлен на email!"})

    return JsonResponse({"success": False, "error": "Неверный метод запроса!"})

@csrf_exempt
def verify_code(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get("email")
        code = int(data.get("code"))

        print(f"Проверяем код для {email}: {code}")

        if email in verification_codes and verification_codes[email] == code:
            del verification_codes[email]  # Удаляем код после успешной верификации
            request.session["verified_email"] = email  # Сохраняем email в сессии
            request.session.modified = True  # Принудительно сохраняем сессию
            print(f"✅ Код верный! Email {email} сохранен в сессии.")
            return JsonResponse({"success": True, "message": "Код подтвержден!"})

        print(f"❌ Код неверный для {email}!")
        return JsonResponse({"success": False, "error": "Неверный код!"})

    return JsonResponse({"error": "Неверный запрос"}, status=400)

def register_form(request):
    # 🔒 Если пользователь уже залогинен — перебросить на страницу тьюторов
    if request.user.is_authenticated:
        return redirect("/tutors/")

    email = request.session.get("verified_email")
    print(f"🔍 [DEBUG] Email в сессии перед загрузкой формы: {email}")

    if not email:
        return redirect("/register/")

    return render(request, "tutor_app/register_form.html", {"email": email})

from django.contrib.auth import login
from django.shortcuts import redirect
import json
from django.http import JsonResponse
from tutor_app.models import CustomUser
from django.views.decorators.csrf import csrf_exempt

from django.contrib.auth import authenticate, login

@csrf_exempt
def complete_registration(request):
    if request.method == "POST":
        data = json.loads(request.body)
        email = data.get("email")
        first_name = data.get("first_name")
        last_name = data.get("last_name")
        group = data.get("group")
        course = data.get("course")
        password = data.get("password")

        try:
            # ✅ Создаем нового пользователя
            user = CustomUser.objects.create_user(email=email, password=password)
            user.first_name = first_name
            user.last_name = last_name
            user.group = group
            user.save()

            # ✅ Автоматический вход после регистрации
            login(request, user)
            print(f"✅ {email} зарегистрирован и вошел!")

        except IntegrityError:
            # ❗️ Если пользователь уже существует, просто входим в аккаунт
            print(f"⚠️ Пользователь {email} уже существует, выполняем вход...")

            user = authenticate(request, username=email, password=password)
            if user is not None:
                login(request, user)
                print(f"✅ {email} успешно вошел!")
            else:
                print(f"❌ Ошибка входа для {email} (возможно, неверный пароль).")
                return JsonResponse({"success": False, "error": "Неверный пароль для существующего аккаунта!"}, status=400)

        return JsonResponse({"success": True})  # ❌ Убираем второй редирект

    return JsonResponse({"success": False, "error": "Неверный метод запроса!"})




@login_required(login_url='/login/')

def edit_tutor_profile(request):
    if not request.user.is_authenticated or not request.user.is_tutor:
        return redirect("tutors")

    profile, created = TutorProfile.objects.get_or_create(user=request.user)

    # Инициализация переменных для GET-запроса
    course = None
    subject = None

    if request.method == "POST":
        form = TutorProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            profile = form.save(commit=False)
            course = request.POST.get('course')  # Получаем курс
            subject = request.POST.get('subject')  # Получаем предмет
            telegram = request.POST.get('telegram')
            phone = request.POST.get('phone')
            print(f"Выбранный курс: {course}")
            print(f"Выбранный предмет: {subject}")
            print(f"телега: {telegram}")

            # Обновляем имя и фамилию из User
            profile.first_name = request.user.first_name
            profile.last_name = request.user.last_name
            profile.telegram = telegram
            profile.phone = phone
            # Декодируем JSON расписания из запроса
            try:
                schedule_data = json.loads(request.POST.get("schedule", "{}"))
                profile.schedule = schedule_data  # Сохраняем в JSONField
            except json.JSONDecodeError:
                return JsonResponse({"success": False, "errors": {"schedule": ["Некорректный формат расписания"]}})

            profile.save()
            return JsonResponse({"success": True})

        return JsonResponse({"success": False, "errors": form.errors})

    else:
        form = TutorProfileForm(instance=profile)

    # Передаем переменные в контекст шаблона
    return render(request, "tutor_app/edit_tutor_profile.html", {
        "form": form,
        "schedule_json": json.dumps(profile.schedule),  # Передаем в JSON
        "course": course,  # Передаем курс
        "subject": subject,  # Передаем предмет
    })

# ПОПАП МЕНЮ ПРОФИЛЯ
def get_registered_tutors(request):
    user_id = request.GET.get("user_id")
    if not user_id:
        return JsonResponse({"error": "User ID required"}, status=400)

    user = CustomUser.objects.get(id=user_id)

    if user.role == "student":
        # ✅ Студент: Возвращаем тьюторов, к которым записан студент
        registered_tutors = TutorStudent.objects.filter(student_id=user_id, confirmed=True).select_related("tutor")

        tutors_list = [
            {
                "name": tutor.tutor.first_name + " " + tutor.tutor.last_name,
                "subject": tutor.tutor.subject
            }
            for tutor in registered_tutors
        ]
        return JsonResponse(tutors_list, safe=False)

    elif user.role == "tutor":
        # ✅ Тьютор: Возвращаем студентов, записанных к нему, и тьюторов, к которым он записан
        students = TutorStudent.objects.filter(tutor__user_id=user_id, confirmed=True).select_related("student")
        registered_tutors = TutorStudent.objects.filter(student_id=user_id, confirmed=True).select_related("tutor")

        students_list = [
            {
                "name": student.student.first_name + " " + student.student.last_name,
                "email": student.student.email
            }
            for student in students
        ]

        tutors_list = [
            {
                "name": tutor.tutor.first_name + " " + tutor.tutor.last_name,
                "subject": tutor.tutor.subject
            }
            for tutor in registered_tutors
        ]

        return JsonResponse({"students": students_list, "tutors": tutors_list}, safe=False)

@login_required
def get_tutor_schedule(request):
    user_id = request.GET.get("user_id")

    try:
        tutor = TutorProfile.objects.get(user_id=user_id)
        return JsonResponse(tutor.schedule, safe=False)
    except TutorProfile.DoesNotExist:
        return JsonResponse({"error": "Тьютор не найден"}, status=404)
    
# РЕГИСТРАЦИЯ ТЬЮТОРОВ
def get_connection():
    return pymysql.connect(
        host=DB_HOST, user=DB_USER, password=DB_PASSWORD, database=DB_NAME, charset="utf8mb4"
    )



# ✅ Страница регистрации
from django.shortcuts import render, get_object_or_404
from .models import TutorInviteToken, TutorProfile, CustomUser

from django.shortcuts import render, get_object_or_404
from .models import TutorInviteToken, TutorProfile, CustomUser

@csrf_exempt
def register_tutor_page(request, token):
    if not token:
        return JsonResponse({"error": "Token is required"}, status=400)

    try:
        conn = get_connection()
        cursor = conn.cursor()

        # Получаем Telegram ID по токену
        cursor.execute("SELECT user_id FROM tutor_app_tokens WHERE token = %s AND status = 'unused'", (token,))
        row = cursor.fetchone()

        if not row:
            return JsonResponse({"error": "Invalid or used token"}, status=400)

        telegram_id = row[0]

        server_ip = get_server_ip()
        confirm_link = f"http://{server_ip}:8000/confirm_registration/{token}/"

        return render(request, "tutor_app/register_tutor.html", {
            "token": token,
            "confirm_link": confirm_link,
            "telegram_id": telegram_id,  # передай в HTML
            "tutor_name": "",            # пока не знаем
            "subject": "",               # пока не знаем
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({"error": f"Ошибка сервера: {str(e)}"}, status=500)








@csrf_exempt
def complete_tutor_registration(request):
    if request.method == "POST":
        try:
            if request.content_type == "application/json":
                data = json.loads(request.body.decode("utf-8"))
            else:
                data = request.POST

            print("📌 [DEBUG] Полученные данные:", data)

            token = data.get("token")
            telegram_id = data.get("telegram_id")
            first_name = data.get("first_name")
            last_name = data.get("last_name")
            group = data.get("group")
            course = data.get("course")
            email = data.get("email")
            password = data.get("password")
            print("📦 telegram_id:", telegram_id)  # ✅ А ТАКЖЕ ЭТО
            if not all([token, telegram_id, first_name, last_name, email, password]):
                return JsonResponse({"error": "Заполните все обязательные поля!"}, status=400)

            if not email.endswith("@auca.kg"):
                return JsonResponse({"error": "Email должен быть на @auca.kg"}, status=400)

            existing_user = CustomUser.objects.filter(email=email, role="tutor").first()
            if existing_user:
                return JsonResponse({"error": "Пользователь с такой почтой уже зарегистрирован как тьютор!"}, status=400)
            
            conn = get_connection()
            cursor = conn.cursor()

            # Проверяем валидность токена
            cursor.execute("SELECT id FROM tutor_app_tokens WHERE token = %s AND status = 'unused'", (token,))
            row = cursor.fetchone()

            if not row:
                return JsonResponse({"error": "Недействительный или использованный токен"}, status=400)

            # Создаём или получаем пользователя
            user, created = CustomUser.objects.get_or_create(email=email)

            user.first_name = first_name
            user.last_name = last_name
            user.telegram_id = telegram_id  # 💥 Вот здесь сохраняем ID
            user.role = "tutor"
            user.set_password(password)
            user.save()

            # Создаём или обновляем профиль тьютора
            profile, _ = TutorProfile.objects.get_or_create(user=user)
            profile.first_name = first_name
            profile.last_name = last_name
            profile.group = group
            profile.own_course = course
            profile.subject = profile.subject or ""
            profile.save()

            # Обновляем токен
            cursor.execute("UPDATE tutor_app_tokens SET status = 'used' WHERE token = %s", (token,))
            conn.commit()
            conn.close()

            login(request, user)

            return JsonResponse({"success": True, "message": "Регистрация прошла успешно!"})

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": f"Ошибка сервера: {str(e)}"}, status=500)

    return JsonResponse({"error": "Неверный метод запроса!"}, status=405)




# BOTTTT
# ✅ API: Проверка токена (бот будет вызывать этот API перед выдачей ссылки)
@csrf_exempt
def validate_token(request):
    token = request.GET.get("token")

    if not token:
        return JsonResponse({"valid": False, "error": "No token provided"}, status=400)

    with connection.cursor() as cursor:
        cursor.execute("SELECT status FROM tutor_app_tutorinvitetoken WHERE token = %s", [token])
        row = cursor.fetchone()

    if row and row[0] == "unused":
        with connection.cursor() as cursor:
            cursor.execute("UPDATE tutor_app_tutorinvitetoken SET status = 'used' WHERE token = %s", [token])
        return JsonResponse({"valid": True})

    return JsonResponse({"valid": False, "error": "Invalid or used token"}, status=400)


# ✅ API: Проверка регистрации пользователя (бот будет проверять по Telegram ID)
@csrf_exempt
def check_user_registration(request):
    """ Проверяем, зарегистрирован ли пользователь по Telegram ID """
    telegram_id = request.GET.get("user_id")

    if not telegram_id:
        return JsonResponse({"error": "User ID is required"}, status=400)

    is_registered = CustomUser.objects.filter(telegram_id=telegram_id).exists()
    return JsonResponse({"registered": is_registered})



# Запись к тьютору
@login_required
def send_tutor_request(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            tutor_profile_id = data.get("tutor_id")

            print("📦 tutor_id из POST:", tutor_profile_id)

            tutor_profile = get_object_or_404(TutorProfile, id=tutor_profile_id)
            tutor_user = tutor_profile.user
            student = request.user

            # 🔥 Добавляем запись в TutorStudent, если её нет
            relation, created = TutorStudent.objects.get_or_create(
                tutor=tutor_profile,
                student=student,
                defaults={"confirmed": False}
            )

            if not created:
                print("⚠️ Запись уже существует:", relation)
                return JsonResponse({"status": "already_requested", "message": "Вы уже отправили запрос!"})

            # 🔔 Создаем уведомление тьютору
            Notification.objects.create(
            user=tutor_user,
            message=f"📩 Новый запрос от {student.first_name} {student.last_name} на предмет «{tutor_profile.subject}». "
                    f"<span data-student-id='{student.id}' data-student-name='{student.first_name} {student.last_name}' "
                    f"data-tutor-id='{tutor_profile.id}' class='inline-approve'>✅</span> "
                    f"<span data-student-id='{student.id}' data-student-name='{student.first_name} {student.last_name}' "
                    f"data-tutor-id='{tutor_profile.id}' class='inline-decline'>❌</span>"
            )


            print("✅ Уведомление отправлено тьютору")
            return JsonResponse({"status": "ok", "message": "Запрос отправлен тьютору!"})

        except Exception as e:
            print("❌ Ошибка внутри POST обработки:", str(e))
            return JsonResponse({"error": "Ошибка обработки запроса", "details": str(e)}, status=500)

    return JsonResponse({"error": "Неверный метод"}, status=405)




def confirm_tutor_request(request, token):
    try:
        invite = TutorInviteToken.objects.get(token=token)
        if invite.is_used:
            return redirect("/tutors?confirmed=already")

        tutor_user = CustomUser.objects.get(email=invite.email)
        tutor_profile = tutor_user.tutor_profile
        student_user = invite.student

        relation = TutorStudent.objects.get(tutor=tutor_profile, student=student_user)
        relation.confirmed = True
        relation.save()

        invite.is_used = True
        invite.save()

        return redirect(f"/tutors?confirmed=1&student_id={student_user.id}&first_name={student_user.first_name}&last_name={student_user.last_name}")

    except Exception as e:
        print("❌ Ошибка при подтверждении:", e)
        return redirect("/tutors?confirmed=error")



@csrf_exempt
@login_required
def get_my_lessons(request):
    tutor_id = request.GET.get("tutor_id")
    if not tutor_id:
        return JsonResponse({"error": "Tutor ID is required"}, status=400)

    try:
        tutor_profile = TutorProfile.objects.get(user__id=tutor_id)
    except TutorProfile.DoesNotExist:
        return JsonResponse({"error": "Tutor not found"}, status=404)

    schedule = tutor_profile.schedule or {}
    current_time = pytz.timezone("Asia/Bishkek").localize(datetime(2025, 3, 31, 8, 30))
    show_fields = False

    for day, pair_ids in schedule.items():
        for pair_num in pair_ids:
            if pair_num not in lesson_times:
                continue
            start_str, _ = lesson_times[pair_num].split("-")
            start_hour, start_min = map(int, start_str.strip().split(":"))
            start_time = current_time.replace(hour=start_hour, minute=start_min, second=0, microsecond=0)
            if current_time < start_time + timedelta(hours=2):
                show_fields = True

    students = TutorStudent.objects.filter(tutor=tutor_profile, attended=True).select_related("student")
    lessons = []
    for s in students:
        student = s.student
        try:
            student_profile = TutorProfile.objects.get(user=student)
            group = student_profile.group or student.group or ""
        except TutorProfile.DoesNotExist:
            group = student.group or ""

        lessons.append({
            "first_name": student.first_name,
            "last_name": student.last_name,
            "group": group
        })

    return JsonResponse({
        "lessons": lessons,
        "show_fields": show_fields,
        "schedule": schedule,  # <--- 🔥 вот эту строку добавь
        "topic": tutor_profile.last_lesson_topic or "",
        "location": tutor_profile.last_lesson_location or ""
    })




def send_acceptance_notification(request):
    if request.method == "POST":
        data = json.loads(request.body)
        student_id = data.get("student_id")
        tutor_name = data.get("tutor_name")
        subject = data.get("subject")
        
        # Получаем студента
        student = CustomUser.objects.get(id=student_id)

        # Создаем уведомление
        notification_message = f"{tutor_name} принял вашу заявку на предмет {subject}."
        notification_time = datetime.now().strftime("%H:%M")

        # Сохраняем уведомление в базе данных
        notification = Notification.objects.create(
            user=student,
            message=f"{notification_message} Время: {notification_time}",
            timestamp=datetime.now()
        )

        return JsonResponse({"success": True, "message": "Уведомление отправлено!"})
    return JsonResponse({"error": "Неверный запрос!"}, status=400)


@login_required(login_url='/login/')
def get_notifications(request):
    user = request.user
    notifications = Notification.objects.filter(user=user).order_by('-timestamp')  # Сортируем по времени

    # Преобразуем уведомления в формат, удобный для фронтенда
    notifications_data = []
    for notification in notifications:
        notifications_data.append({
            "message": notification.message,
            "timestamp": notification.timestamp.isoformat(),
            "is_read": notification.is_read
        })
    
    return JsonResponse({"notifications": notifications_data})


@csrf_exempt
@login_required
def confirm_via_notification(request):
    data = json.loads(request.body)
    student_id = data.get("student_id")
    tutor_id = data.get("tutor_id")

    relation = TutorStudent.objects.get(tutor_id=tutor_id, student_id=student_id)
    relation.confirmed = True
    relation.save()

    student = relation.student
    tutor_profile = relation.tutor
    tutor_user = tutor_profile.user

    # ✅ Уведомление студенту
    Notification.objects.create(
        user_id=student.id,
        message=(
            f"🎉 Ваша заявка к тьютору {tutor_user.first_name} {tutor_user.last_name} "
            f"по предмету «{tutor_profile.subject}» принята!"
        )
    )

    # ✅ История действия для тьютора
    Notification.objects.create(
        user=request.user,
        message=(
            f"✅ Вы приняли заявку от {student.first_name} {student.last_name} "
        )
    )

    # ❌ Удаляем исходное уведомление с кнопками
    Notification.objects.filter(
        user=request.user,
        message__icontains=f"Новый запрос от {student.first_name} {student.last_name}"
    ).delete()

    return JsonResponse({"success": True})



@csrf_exempt
@login_required
def decline_tutor_request(request):
    data = json.loads(request.body)
    student_id = data.get("student_id")
    tutor_id = data.get("tutor_id")

    try:
        student = CustomUser.objects.get(id=student_id)
        tutor_profile = TutorProfile.objects.get(id=tutor_id)
        tutor_user = request.user

        # Удаляем связь
        TutorStudent.objects.filter(tutor_id=tutor_id, student_id=student_id).delete()

        # ✅ Уведомление студенту об отказе
        Notification.objects.create(
            user_id=student.id,
            message=(
                f"❌ Ваша заявка к тьютору {tutor_user.first_name} {tutor_user.last_name} "
                f"по предмету «{tutor_profile.subject}» была отклонена."
            )
        )

        # ✅ Уведомление тьютору (история отказа)
        Notification.objects.create(
            user=request.user,
            message=(
                f"❌ Вы отклонили заявку от {student.first_name} {student.last_name} "
            )
        )

        # ❌ Удаляем исходное уведомление с кнопками
        Notification.objects.filter(
            user=request.user,
            message__icontains=f"Новый запрос от {student.first_name} {student.last_name}"
        ).delete()

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)
    




@csrf_exempt
@login_required
def get_tutor_schedule_with_attendance(request):
    try:
        user = request.user
        tutor_students = TutorStudent.objects.filter(student=user, confirmed=True).select_related("tutor", "tutor__user")

        if not tutor_students.exists():
            return JsonResponse({"error": "У вас нет ни одного тьютора"}, status=403)

        result = []
        # current_time = datetime.now(pytz.timezone("Asia/Bishkek"))  # в проде
        current_time = pytz.timezone("Asia/Bishkek").localize(datetime(2025, 3, 31, 8, 30))
        current_day = current_time.strftime("%A")  # English: Monday, etc.
        print(current_time)
        print(current_day)
        # Карта перевода на русский
        day_map = {
            "Monday": "Понедельник",
            "Tuesday": "Вторник",
            "Wednesday": "Среда",
            "Thursday": "Четверг",
            "Friday": "Пятница",
            "Saturday": "Суббота",
            "Sunday": "Воскресенье"
        }
        current_day_rus = day_map.get(current_day)

        for relation in tutor_students:
            tutor = relation.tutor
            schedule = tutor.schedule or {}

            formatted_schedule = []

            for day, pair_ids in schedule.items():
                for pair_num in pair_ids:
                    if pair_num not in lesson_times:
                        continue

                    time_range = lesson_times[pair_num]
                    start_str, _ = time_range.split("-")
                    start_hour, start_min = map(int, start_str.strip().split(":"))
                    start_time = current_time.replace(hour=start_hour, minute=start_min, second=0, microsecond=0)

                    # ✅ Проверка на отметку
                    allowed_to_mark = (
                        day.lower() == current_day_rus.lower() and
                        current_time >= start_time and
                        current_time <= (start_time + timedelta(hours=2))
                    )

                    # ✅ Автоэкспорт Excel, если прошло 2 часа с начала пары
                    if not tutor.excel_generated and current_time > start_time + timedelta(hours=2):
                        generate_excel_for_lesson(tutor)
                        tutor.excel_generated = True
                        tutor.save()
                        print(f"💾 Excel создан для тьютора {tutor.user.get_full_name()}")

                    formatted_schedule.append({
                        "day": day,
                        "time": time_range,
                        "can_mark": allowed_to_mark
                    })

            result.append({
                "tutor_id": tutor.id,
                "tutor_name": f"{tutor.user.first_name} {tutor.user.last_name}",
                "subject": tutor.subject,
                "schedule": formatted_schedule,
                "already_marked": relation.attended
            })

        return JsonResponse({"schedule": result})

    except Exception as e:
        print("🔥 Ошибка в get_schedule_with_attendance:", e)
        import traceback
        traceback.print_exc()
        return JsonResponse({"error": "Внутренняя ошибка сервера"}, status=500)





@csrf_exempt
@login_required
def submit_attendance(request):
    data = json.loads(request.body)
    tutor_id = data.get("tutor_id")
    user = request.user

    try:
        relation = TutorStudent.objects.get(tutor_id=tutor_id, student=user, confirmed=True)
        tutor = relation.tutor
        schedule = tutor.schedule or {}

        # Используем фиксированное тестовое время (или datetime.now(...) в проде)
        current_time = pytz.timezone("Asia/Bishkek").localize(datetime(2025, 3, 31, 8, 30))
        current_day = current_time.strftime("%A")  # "Monday"

        # Переводим в русский, как в расписании
        day_map = {
            "Monday": "Понедельник",
            "Tuesday": "Вторник",
            "Wednesday": "Среда",
            "Thursday": "Четверг",
            "Friday": "Пятница",
            "Saturday": "Суббота",
            "Sunday": "Воскресенье"
        }
        current_day_rus = day_map.get(current_day)

        lesson_times = {
            "1": "08:00-09:20",
            "2": "09:30-10:50",
            "3": "11:40-13:00",
            "4": "13:10-14:30",
            "5": "14:40-16:00",
            "6": "16:00-17:20",
            "7": "17:30-19:00"
        }

        allowed = False

        if current_day_rus in schedule:
            for pair_id in schedule[current_day_rus]:
                if pair_id not in lesson_times:
                    continue
                start_str, _ = lesson_times[pair_id].split("-")
                start_hour, start_min = map(int, start_str.strip().split(":"))
                start_time = current_time.replace(hour=start_hour, minute=start_min, second=0, microsecond=0)
                if current_time >= start_time and current_time <= (start_time + timedelta(hours=2)):
                    allowed = True
                    break

        if not allowed:
            return JsonResponse({"error": "Время отметки вышло"}, status=403)

        # ✅ Сохраняем факт отметки
        relation.attended = True
        relation.attendance_time = current_time
        relation.save()

        return JsonResponse({"success": True, "message": "Посещение отмечено!"})
    
    except TutorStudent.DoesNotExist:
        return JsonResponse({"error": "Запись не найдена"}, status=404)




@csrf_exempt
@login_required
def save_lesson_details(request):
    data = json.loads(request.body)
    tutor_id = data.get("tutor_id")
    topic = data.get("topic")
    location = data.get("location")

    if not tutor_id or not topic or not location:
        return JsonResponse({"error": "Все поля обязательны"}, status=400)

    try:
        tutor_profile = TutorProfile.objects.get(user_id=tutor_id)
        tutor_profile.last_lesson_topic = topic
        tutor_profile.last_lesson_location = location
        tutor_profile.save()
        return JsonResponse({"success": True})
    except TutorProfile.DoesNotExist:
        return JsonResponse({"error": "Тьютор не найден"}, status=404)
    

def generate_excel_for_lesson(tutor: TutorProfile):
    students = TutorStudent.objects.filter(tutor=tutor, attended=True).select_related("student")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Урок"

    # Шапка
    ws["A1"] = f"Тема: {tutor.last_lesson_topic or 'не указана'}"
    ws["A2"] = f"Место: {tutor.last_lesson_location or 'не указано'}"
    ws["A3"] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws.append([])

    # Заголовки таблицы
    ws.append(["Фамилия", "Имя", "Группа"])
    for col in ws[5]:
        col.font = Font(bold=True)

    # Список студентов
    for s in students:
        student = s.student
        try:
            profile = TutorProfile.objects.get(user=student)
            group = profile.group or student.group or ""
        except TutorProfile.DoesNotExist:
            group = student.group or ""
        ws.append([student.last_name, student.first_name, group])

    # Путь до media/lessons/
    folder_path = os.path.join(settings.MEDIA_ROOT, "lessons")
    os.makedirs(folder_path, exist_ok=True)

    file_name = f"lesson_{tutor.user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(folder_path, file_name)

    wb.save(file_path)

    print(f"✅ Урок экспортирован в Excel: {file_name}")
    return file_name